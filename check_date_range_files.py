"""
One-time script to check files from FTP accounts for date range: Oct 16, 2025 to Oct 22, 2025
This script creates date-wise Excel files in horizontal format matching the existing template.
"""

import datetime
import logging
from pathlib import Path
import sys
from stat import S_ISREG

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from src.config import load_multiple_configs_from_file
from src.sftp_connector import SFTPConnector


# Setup directories
if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent

RESULT_DIR = BASE_DIR / "result"
ERROR_DIR = BASE_DIR / "errors"
LOG_DIR = BASE_DIR / "logs"

# Create directories if they don't exist
RESULT_DIR.mkdir(exist_ok=True)
ERROR_DIR.mkdir(exist_ok=True)
LOG_DIR.mkdir(exist_ok=True)


def setup_logging():
    """Setup logging configuration."""
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = LOG_DIR / f"date_range_check_{timestamp}.log"
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)


def check_files_for_date_range(start_date: datetime.date, end_date: datetime.date):
    """
    Check files from FTP accounts for each date in the specified range.
    
    Args:
        start_date: Starting date (inclusive)
        end_date: Ending date (inclusive)
    """
    logger = logging.getLogger(__name__)
    configs = load_multiple_configs_from_file()
    all_results = []
    
    # Generate list of dates in range
    date_range = []
    current_date = start_date
    while current_date <= end_date:
        date_range.append(current_date)
        current_date += datetime.timedelta(days=1)
    
    logger.info(f"Checking files from {start_date} to {end_date} ({len(date_range)} days)")
    
    for config in configs:
        account_name = config["name"]
        logger.info(f"\n{'='*80}")
        logger.info(f"Processing Account: {account_name}")
        logger.info(f"Host: {config.get('host', '-')}")
        logger.info(f"{'='*80}")
        
        connector = SFTPConnector(
            host=config["host"],
            port=config.get("port", 22),
            username=config["username"],
            password=config["password"],
        )
        
        try:
            connector.connect()
            logger.info(f"[OK] Connected to '{account_name}'")
        except Exception as exc:
            logger.error(f"[ERROR] Connection error for '{account_name}': {exc}")
            for date in date_range:
                all_results.append({
                    "Account Name": account_name,
                    "Folder": "-",
                    "Date": date.strftime('%Y-%m-%d'),
                    "File Name": "-",
                    "LastFileDate": f"Connection error: {exc}",
                })
            connector.disconnect()
            continue
        
        try:
            folders = config.get("folders", [])
            
            if not folders:
                logger.warning(f"No folders configured for '{account_name}'")
                for date in date_range:
                    all_results.append({
                        "Account Name": account_name,
                        "Folder": "-",
                        "Date": date.strftime('%Y-%m-%d'),
                        "File Name": "-",
                        "File DateTime": "No folders configured",
                    })
                continue
            
            for folder in folders:
                folder_label = folder.get("label", "Folder")
                folder_path = folder.get("path", "/")
                filters = folder.get("filters", [])
                
                logger.info(f"\nChecking folder: '{folder_label}' ({folder_path})")
                if filters:
                    logger.info(f"  Filters: {filters}")
                
                # Check files for each date in the range
                for check_date in date_range:
                    try:
                        if filters:
                            # For folders with multiple prefixes (e.g., Inventory)
                            prefix_results = connector.get_latest_files_per_prefix(
                                folder_path, 
                                filters, 
                                on_date=check_date
                            )
                            
                            for prefix, (filename, file_dt) in prefix_results.items():
                                result = {
                                    "Account Name": account_name,
                                    "Folder": f"{folder_label} ({prefix})",
                                    "Date": check_date.strftime('%Y-%m-%d'),
                                    "File Name": filename if filename else "No file found",
                                    "LastFileDate": file_dt if file_dt else "-",
                                }
                                all_results.append(result)
                                
                                if filename:
                                    logger.info(f"  [{check_date}] {prefix}: {filename} - {file_dt}")
                                else:
                                    logger.info(f"  [{check_date}] {prefix}: No file found")
                        else:
                            # For folders without specific filters
                            filename, file_dt = connector.get_latest_file_info_on_date(
                                folder_path,
                                on_date=check_date,
                                name_filters=None
                            )
                            
                            result = {
                                "Account Name": account_name,
                                "Folder": folder_label,
                                "Date": check_date.strftime('%Y-%m-%d'),
                                "File Name": filename if filename else "No file found",
                                "LastFileDate": file_dt if file_dt else "-",
                            }
                            all_results.append(result)
                            
                            if filename:
                                logger.info(f"  [{check_date}] {filename} - {file_dt}")
                            else:
                                logger.info(f"  [{check_date}] No file found")
                                
                    except Exception as exc:
                        logger.error(f"  [{check_date}] Error checking folder '{folder_label}': {exc}")
                        all_results.append({
                            "Account Name": account_name,
                            "Folder": folder_label,
                            "Date": check_date.strftime('%Y-%m-%d'),
                            "File Name": "-",
                            "LastFileDate": f"Error: {exc}",
                        })
        
        except Exception as exc:
            logger.error(f"Error processing account '{account_name}': {exc}")
        finally:
            connector.disconnect()
            logger.info(f"[OK] Disconnected from '{account_name}'")
    
    return all_results


def create_horizontal_format_excel(check_date, results, output_file):
    """Create Excel file in horizontal format matching the template."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create worksheet")
    ws.title = "Daily Update"
    
    # Group results by account
    account_data = {}
    for result in results:
        if result['Date'] == check_date.strftime('%Y-%m-%d'):
            account = result['Account Name']
            if account not in account_data:
                account_data[account] = []
            account_data[account].append(result)
    
    current_row = 1
    
    # Define styles
    header_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    
    for account_name, account_results in account_data.items():
        # Account header row
        ws.cell(row=current_row, column=4, value=account_name).font = header_font
        ws.cell(row=current_row, column=5, value='LastFileDate').font = header_font
        current_row += 1
        
        # Add each folder's data
        for result in account_results:
            folder_label = result['Folder']
            file_date = result['LastFileDate']
            
            # Column B: Account Name (for some rows)
            if result['Account Name']:
                ws.cell(row=current_row, column=2, value=result['Account Name'])
            
            # Column C: Folder name
            ws.cell(row=current_row, column=3, value=folder_label)
            
            # Column D: Folder type/category
            folder_type = folder_label.split('(')[0].strip() if '(' not in folder_label else folder_label
            ws.cell(row=current_row, column=4, value=folder_type)
            
            # Column E: LastFileDate
            if file_date != '-' and file_date != 'No file found' and 'error' not in file_date.lower():
                # Parse and format date
                try:
                    dt = datetime.datetime.strptime(file_date, '%m/%d/%Y %H:%M:%S')
                    ws.cell(row=current_row, column=5, value=dt)
                except:
                    ws.cell(row=current_row, column=5, value=file_date)
            else:
                ws.cell(row=current_row, column=5, value=file_date)
            
            current_row += 1
        
        # Add blank rows between accounts
        current_row += 2
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    
    wb.save(output_file)


def main():
    """Main function to run the date range check."""
    logger = setup_logging()
    
    # Define date range: Oct 16, 2025 to Oct 22, 2025 (today)
    start_date = datetime.date(2025, 10, 16)
    end_date = datetime.date(2025, 10, 22)
    
    logger.info("="*80)
    logger.info("FTP Files Date Range Check - Horizontal Format")
    logger.info(f"Date Range: {start_date} to {end_date}")
    logger.info("="*80)
    
    try:
        results = check_files_for_date_range(start_date, end_date)
        
        if results:
            # Create DataFrame for reference
            df = pd.DataFrame(results)
            
            # Generate list of dates in range
            date_range = []
            current_date = start_date
            while current_date <= end_date:
                date_range.append(current_date)
                current_date += datetime.timedelta(days=1)
            
            # Create separate Excel files for each date in horizontal format
            logger.info("\nCreating date-wise Excel files in horizontal format...")
            files_created = []
            
            for check_date in date_range:
                # Create filename in format: Accounts_Daily_Update MM-DD-YYYY.xlsx (with space)
                output_filename = f"Accounts_Daily_Update {check_date.strftime('%m-%d-%Y')}.xlsx"
                output_file = RESULT_DIR / output_filename
                
                # Create the horizontal format Excel
                create_horizontal_format_excel(check_date, results, output_file)
                
                # Count files for this date
                date_results = [r for r in results if r['Date'] == check_date.strftime('%Y-%m-%d')]
                found_count = sum(1 for r in date_results if r['LastFileDate'] not in ['-', 'No file found'] and 'error' not in r['LastFileDate'].lower())
                missing_count = sum(1 for r in date_results if r['LastFileDate'] == '-' or r['File Name'] == 'No file found')
                error_count = sum(1 for r in date_results if 'error' in r['LastFileDate'].lower())
                
                files_created.append((output_filename, found_count, missing_count, error_count))
                logger.info(f"  [OK] {output_filename} - Found: {found_count}, Missing: {missing_count}, Errors: {error_count}")
            
            # Print summary
            logger.info(f"\n{'='*80}")
            logger.info("SUMMARY")
            logger.info(f"{'='*80}")
            logger.info(f"Total accounts checked: {df['Account Name'].nunique()}")
            logger.info(f"Total folders checked: {df['Folder'].nunique()}")
            logger.info(f"Files found: {sum(1 for r in results if r['LastFileDate'] not in ['-', 'No file found'] and 'error' not in r['LastFileDate'].lower())}")
            logger.info(f"Files missing: {sum(1 for r in results if r['File Name'] == 'No file found')}")
            logger.info(f"\nDate-wise Excel files created: {len(files_created)}")
            for filename, found, missing, errors in files_created:
                logger.info(f"  - {filename}: Found={found}, Missing={missing}, Errors={errors}")
            logger.info(f"{'='*80}")
            
        else:
            logger.warning("No results collected.")
    
    except Exception as exc:
        logger.error(f"Error during execution: {exc}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
